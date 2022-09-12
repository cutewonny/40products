from openpyxl import load_workbook

load_wb = load_workbook(r'/Users/devwonny/Documents/40products/12. 엑셀의 정보를 불러와 수료증 자동 생성/엑셀생성.xlsx')
load_ws = load_wb.active

name_list = [] #1번 항목 : 이름
birthday_list = [] #2번 항목: 생일
ho_list = [] #3번 항목: 사원번호

print(load_ws.max_row) # 3개의 row
for i in range(1, load_ws.max_row +1):
    name_list.append(load_ws.cell(i,1).value)
    birthday_list.append(load_ws.cell(i,2).value)
    ho_list.append(load_ws.cell(i,3).value)

print(name_list)
print(birthday_list)
print(ho_list)